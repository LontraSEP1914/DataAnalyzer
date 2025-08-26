# core/report_generator.py

import pandas as pd
import numpy as np # Para np.inf
import os
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Para usar as funções dos nossos outros módulos para teste
try:
    from .excel_parser import carregar_dados_excel
    from .data_comparator import comparar_dataframes # Importado para o teste __main__
except ImportError:
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    if parent_dir not in sys.path:
        sys.path.append(parent_dir)
    from core.excel_parser import carregar_dados_excel
    from core.data_comparator import comparar_dataframes


def aplicar_estilos_planilha(writer, sheet_name, df_para_estilo):
    """Aplica formatação e estilos a uma planilha."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # --- Configurações da Planilha ---
    worksheet.sheet_view.showGridLines = False  # Remove linhas de grade
    worksheet.sheet_view.zoomScale = 85         # Define zoom para 85%

    # --- Estilos ---
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Preto
    header_font = Font(bold=True, color="FFFFFF") # Texto branco
    
    thin_border_side = Side(border_style="thin", color="D0D0D0") # Cinza claro para bordas
    thin_border = Border(left=thin_border_side, 
                         right=thin_border_side, 
                         top=thin_border_side, 
                         bottom=thin_border_side)

    # Aplicar ao Cabeçalho
    for col_num, column_title in enumerate(df_para_estilo.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Adicionado wrap_text
        cell.border = thin_border


    # Ajustar Largura das Colunas e Aplicar Bordas e Formatos às Células de Dados
    for r_idx in range(2, worksheet.max_row + 1): # Começa da linha 2 (dados)
        for c_idx, col_name in enumerate(df_para_estilo.columns, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.border = thin_border # Aplicar borda fina a todas as células de dados

            # Lógica de Formatação de Número (mantida e ajustada)
            if sheet_name == "Resumo_Comparacao":
                if col_name == 'Diferença Percentual Total (%)':
                    cell.number_format = '0.00%'
                elif col_name in ['Total Lado A', 'Total Lado B', 'Diferença Absoluta Total']:
                    cell.number_format = '#,##0.00'
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            elif sheet_name == "Dados_Detalhados":
                if col_name.endswith("_DiffPerc_Linha(%)"):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif cell.value in ['INF', '-INF']: # Strings 'INF'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name.endswith("_DiffAbs_Linha") or \
                     (col_name.endswith("_A") and not col_name.endswith(("_original_A", "_chave_original_A"))) or \
                     (col_name.endswith("_B") and not col_name.endswith(("_original_B", "_chave_original_B"))):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                # Alinhar texto à esquerda por padrão para outras colunas
                elif isinstance(cell.value, str) and cell.value not in ['INF', '-INF']:
                     cell.alignment = Alignment(horizontal="left", vertical="center")


    # Ajustar Largura das Colunas (após preencher e formatar)
    for col_num, column_title in enumerate(df_para_estilo.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        # Cabeçalho
        if column_title:
            max_length = len(str(column_title))

        # Dados
        if column_title in df_para_estilo:
            # Tentar encontrar o comprimento máximo dos dados formatados (como string)
            # Isso é um pouco heurístico pois o number_format pode mudar a largura visual
            # Usar uma combinação do valor original e um fator para formatos comuns
            for i in range(len(df_para_estilo[column_title])):
                cell_value = worksheet.cell(row=i+2, column=col_num).value # Valor como está na célula (pode ser formatado)
                # Se for string 'INF' ou '-INF', usar seu comprimento.
                # Se for número, estimar com base no formato. Por simplicidade, usar str(valor)
                # mas com cuidado para formatos de data/hora longos.
                value_len = len(str(cell_value))
                if worksheet.cell(row=i+2, column=col_num).number_format == '0.00%':
                    value_len = max(value_len, 6) # Espaço para "XX.XX%"
                elif '#' in worksheet.cell(row=i+2, column=col_num).number_format:
                    value_len = max(value_len, 8) # Espaço para números formatados
                max_length = max(max_length, value_len)
        
        adjusted_width = (max_length + 2) * 1.1 # Reduzido um pouco o multiplicador de padding
        worksheet.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 55) # Min 10, Max 55

    # Congelar Painel do Cabeçalho
    worksheet.freeze_panes = worksheet['A2']


def gerar_relatorio_excel(dados_comparacao: dict, caminho_saida: str, nome_planilha_resumo: str = "Resumo_Comparacao", nome_planilha_detalhes: str = "Dados_Detalhados"):
    if not dados_comparacao or 'resumo_por_par' not in dados_comparacao or 'dataframe_merged' not in dados_comparacao:
        print("Erro: Dados de comparação ('resumo_por_par' ou 'dataframe_merged') ausentes.")
        return False
    try:
        lista_resumo_original = dados_comparacao['resumo_por_par']
        df_resumo_para_escrita = pd.DataFrame()
        if not lista_resumo_original:
            df_resumo_para_escrita = pd.DataFrame([{"Status": "Nenhum par de colunas mapeado ou comparável encontrado."}])
        else:
            lista_resumo_modificada = []
            for item_resumo in lista_resumo_original:
                item_copiado = item_resumo.copy()
                if 'diferenca_percentual_total' in item_copiado and isinstance(item_copiado['diferenca_percentual_total'], (int, float)):
                    # Tratar np.inf antes da divisão
                    if item_copiado['diferenca_percentual_total'] == np.inf:
                        item_copiado['diferenca_percentual_total'] = 'INF'
                    elif item_copiado['diferenca_percentual_total'] == -np.inf:
                        item_copiado['diferenca_percentual_total'] = '-INF'
                    elif pd.notna(item_copiado['diferenca_percentual_total']):
                        item_copiado['diferenca_percentual_total'] /= 100.0
                lista_resumo_modificada.append(item_copiado)
            
            df_resumo_para_escrita = pd.DataFrame(lista_resumo_modificada)
            df_resumo_para_escrita.rename(columns={
                'par_comparado': 'Par Comparado', 'total_lado_a': 'Total Lado A',
                'total_lado_b': 'Total Lado B', 'diferenca_absoluta_total': 'Diferença Absoluta Total',
                'diferenca_percentual_total': 'Diferença Percentual Total (%)'
            }, inplace=True)
        
        df_detalhes_original = dados_comparacao['dataframe_merged']
        df_detalhes_para_escrita = df_detalhes_original.copy()

        for col_name in df_detalhes_para_escrita.columns:
            if col_name.endswith("_DiffPerc_Linha(%)"):
                # A coluna já contém números (float) ou np.inf do data_comparator
                # Vamos criar uma nova série para modificação
                col_data = df_detalhes_para_escrita[col_name].copy()
                
                # Identificar onde estão os infinitos
                is_inf = (col_data == np.inf)
                is_neg_inf = (col_data == -np.inf)
                is_finite = np.isfinite(col_data) # Onde não é inf nem NaN

                # Dividir apenas os finitos por 100
                col_data[is_finite] = col_data[is_finite] / 100.0
                
                # Se há infinitos ou se vamos misturar strings, converter para object
                if is_inf.any() or is_neg_inf.any():
                    col_data = col_data.astype(object) # Converte para object para aceitar strings
                    col_data[is_inf] = 'INF'
                    col_data[is_neg_inf] = '-INF'
                
                df_detalhes_para_escrita[col_name] = col_data


        diretorio_saida = os.path.dirname(caminho_saida)
        if diretorio_saida and not os.path.exists(diretorio_saida): os.makedirs(diretorio_saida)

        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_resumo_para_escrita.to_excel(writer, sheet_name=nome_planilha_resumo, index=False)
            aplicar_estilos_planilha(writer, nome_planilha_resumo, df_resumo_para_escrita) 

            df_detalhes_para_escrita.to_excel(writer, sheet_name=nome_planilha_detalhes, index=False)
            aplicar_estilos_planilha(writer, nome_planilha_detalhes, df_detalhes_para_escrita) 
        
        print(f"Relatório gerado com sucesso em: {caminho_saida}")
        return True
    except Exception as e:
        print(f"Ocorreu um erro ao gerar o relatório Excel: {e}")
        import traceback; traceback.print_exc() 
        return False

if __name__ == '__main__':
    print("--- Testando Gerador de Relatório para Mapeamento Explícito de Pares (Dados Simulados Corrigidos) ---")
    
    # Setup de diretórios (mantido como antes)
    current_file_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(current_file_dir) 
    exemplos_dir = os.path.join(base_dir, 'exemplos_excel')
    relatorios_dir = os.path.join(base_dir, 'relatorios_gerados')
    if not os.path.exists(exemplos_dir): os.makedirs(exemplos_dir)
    if not os.path.exists(relatorios_dir): os.makedirs(relatorios_dir)

    # Dados de entrada simulados (para clareza, não são usados diretamente para gerar df_merged_simulado)
    # df_a_sim = pd.DataFrame({
    #     'ID_Transacao': [1, 2, 3, 4], 
    #     'Valor_Venda_Bruta': [10.0, 15.0, 5.0, 7.5], 
    #     'Qtd_Vendida': [10, 15, 20, 5]
    # })
    # df_b_sim = pd.DataFrame({
    #     'Ref': [1, 2, 3, 5], 
    #     'Total_Recebido': [10.0, 14.0, 5.0, 8.0], 
    #     'Unidades': [10, 14, 21, 8]
    # })

    # df_merged_simulado DEVE refletir o que data_comparator.py PRODUZIRIA
    # com um outer join e os cálculos de diferença aplicados.
    
    # Linha 1 (ID 1): A=10, B=10 -> Abs=0, Perc=0%
    # Linha 2 (ID 2): A=15, B=14 -> Abs=1, Perc=(1/14)*100 = 7.14%
    # Linha 3 (ID 3): A=5,  B=5  -> Abs=0, Perc=0%
    # Linha 4 (ID 4, só em A): A=7.5, B=NaN -> Abs=7.5, Perc=INF
    # Linha 5 (ID 5, só em B): A=NaN, B=8   -> Abs=-8,  Perc=(-8/8)*100 = -100%

    df_merged_simulado = pd.DataFrame({
        'ID_Transacao': [1, 2, 3, 4, np.nan], # Chave A
        'Ref': [1, 2, 3, np.nan, 5],          # Chave B
        # Colunas de valor originais (renomeadas como data_comparator faria)
        'Valor_Venda_Bruta_A': [10.0, 15.0, 5.0, 7.5, np.nan],
        'Total_Recebido_B': [10.0, 14.0, 5.0, np.nan, 8.0],
        'Qtd_Vendida_A': [10, 15, 20, 5, np.nan],
        'Unidades_B': [10, 14, 21, np.nan, 8],
        
        # Colunas de diferença como data_comparator as calcularia
        'Valor_Venda_Bruta_vs_Total_Recebido_DiffAbs_Linha': [0.0, 1.0, 0.0, 7.5, -8.0], # CORRIGIDO
        'Valor_Venda_Bruta_vs_Total_Recebido_DiffPerc_Linha(%)': [0.0, 7.14, 0.0, np.inf, -100.0], # CORRIGIDO
        
        'Qtd_Vendida_vs_Unidades_DiffAbs_Linha': [0, 1, -1, 5, -8], # CORRIGIDO
        'Qtd_Vendida_vs_Unidades_DiffPerc_Linha(%)': [0.0, 7.14, -4.76, np.inf, -100.0] # CORRIGIDO
    })

    resumo_por_par_simulado = [
        {
            'par_comparado': 'Valor_Venda_Bruta (A) vs Total_Recebido (B)',
            'total_lado_a': 22.5, # Soma de Valor_Venda_Bruta_A (10+15+5+7.5, NaN tratado como 0)
            'total_lado_b': 27.0, # Soma de Total_Recebido_B (10+14+5+8, NaN tratado como 0)
            'diferenca_absoluta_total': -4.5, 
            'diferenca_percentual_total': (-4.5 / 27.0) * 100 if 27.0 else np.inf # -16.66...
        },
        {
            'par_comparado': 'Qtd_Vendida (A) vs Unidades (B)',
            'total_lado_a': 50, # 10+15+20+5
            'total_lado_b': 53, # 10+14+21+8
            'diferenca_absoluta_total': -3, 
            'diferenca_percentual_total': (-3 / 53.0) * 100 if 53.0 else np.inf # -5.66...
        }
    ]
    # Arredondar percentuais no resumo para consistência
    for item in resumo_por_par_simulado:
        item['diferenca_percentual_total'] = round(item['diferenca_percentual_total'], 2) if pd.notna(item['diferenca_percentual_total']) else item['diferenca_percentual_total']


    resultados_completos_simulados = {
        'resumo_por_par': resumo_por_par_simulado,
        'dataframe_merged': df_merged_simulado
    }

    caminho_relatorio_teste = os.path.join(relatorios_dir, 'relatorio_mapeado_corrigido_teste.xlsx')
    sucesso = gerar_relatorio_excel(resultados_completos_simulados, caminho_relatorio_teste)
    
    if sucesso:
        print(f"Verifique o arquivo: {os.path.abspath(caminho_relatorio_teste)}")
        # print("Este relatório DEVE agora mostrar valores (como 7.50, INF, -100.00%) nas células de diferença, e não brancos.")
    else:
        print("Falha ao gerar o relatório de teste (com dados simulados corrigidos).")